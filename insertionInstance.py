import pandas as pd
from openpyxl import load_workbook


def insertionOfInstanceCount(result_dict):
  excel_file_path = '/home/tspl/Documents/wsrAutomation/acp.xlsx'
  sheet_name = 'Instance Count'
  df = pd.read_excel(excel_file_path, sheet_name=sheet_name)

  df.loc[df['Priority'] == 'P1', 'Mirror'] = result_dict.get('p1', 0)
  df.loc[df['Priority'] == 'p2', 'Mirror'] = result_dict.get('p2', 0)
  df.loc[df['Priority'] == 'p3', 'Mirror'] = result_dict.get('p3', 0)
  df.loc[df['Priority'] == 'P4', 'Mirror'] = result_dict.get('p4', 0)
  df.at[4, 'Mirror'] = result_dict.get('total', 0)

  wb = load_workbook(excel_file_path)
  if sheet_name not in wb.sheetnames:
      wb.create_sheet(title=sheet_name)

  ws = wb[sheet_name]
  for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
      for cell in row:
          cell.value = None

  for row_index, row in enumerate(df.iterrows(), start=2):
      for col_index, (header, value) in enumerate(row[1].items(), start=1):
          ws.cell(row=row_index, column=col_index, value=value)

  wb.save(excel_file_path)

